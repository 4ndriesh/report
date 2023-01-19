# -*- coding: utf-8 -*-
# Чтение файла шаблона xlsx для получение параметров
import os
import shutil

from functools import wraps
import pandas as pd
from openpyxl import load_workbook

from singleton import singleton

@singleton
class Xlsx:
    def __init__(self):
        pass
    def validate_args(func):
        @wraps(func)
        def wrapped(*args, **kwargs):
            path = kwargs.get('path')
            if (os.path.exists(path)):
                print(path)
                return func(*args, **kwargs)
            else:
                print('{0}{1}'.format('Проверить ', path))
                return pd.DataFrame()
                # raise Exception('{0}{1}'.format('Проверить ', path))
        return wrapped

    @validate_args
    def open_sheet(self, sheet_name=None, skiprows=7,path=''):
        dict_Sheet=pd.read_excel(path, sheet_name=sheet_name, skiprows=skiprows)
        return dict_Sheet

    @validate_args
    def open_book(self,path=''):
        wb=load_workbook(path, read_only=True)
        listsheets = wb.sheetnames
        wb.close()
        return listsheets

    def copy_xlsx(self,path_temp,path_temp_to):
        shutil.copy(path_temp, path_temp_to)

    @validate_args
    def read_csv(self,path='',skiprows=1,delimiter=';'):
            return pd.read_csv(path, encoding='cp1251', delimiter=delimiter, skiprows=skiprows)

    @validate_args
    def write_to_excel(self, data_for_record,startrow, index, conv_file,path=''):
        mv = load_workbook(path)
        with pd.ExcelWriter(path, engine='openpyxl') as write_to_report:
            write_to_report.book = mv
            write_to_report.sheets = dict((ws.title, ws) for ws in mv.worksheets)
            for Sheet in data_for_record.keys():
                data_for_record.get(Sheet).to_excel(write_to_report,
                                                         sheet_name=Sheet,
                                                         startrow=startrow,
                                                         startcol=0,
                                                         header=False,
                                                         index=index)
                sh = mv[Sheet]
                sh.cell(row=4, column=1, value='{0}{1}'.format('Станция ', conv_file))
            write_to_report.save()