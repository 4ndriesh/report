import re
import os
from openpyxl import load_workbook, Workbook
# from ini_file import open_ini_file

dir_filters = {
    'project': re.compile('.*(\.xlsx)$'),
    '3': re.compile('.*(\.xlsx)$'),
    '44': re.compile('.*(\.xlsx)$')
    # 'report': re.compile('.*(\.xlsx)$')
}

# ini = open_ini_file.inst()

def get_list_project_with_value():
    list_station = []
    file_project = []
    list_report = []
    for dir, list_dir, file in walk_dbsta(ini.BASE_DIR):
        if dir == 'project':
            file_project.append(file)
            list_station.append(get_list_sheet(os.path.join(list_dir, file)))
        if dir == 'report':
            list_report.append(file)
    # sorted(list_report)
    s = []
    t = []
    for list_station_in_project in list_station:
        tmp = []
        for station_in_list in list_station_in_project:
            tmp.append(sorted(list(filter(lambda x: x.startswith(station_in_list, 10), list_report))))
        s.append(tmp)

    for i, list_station_in_project in enumerate(list_station):
        t.append(dict(zip(list_station_in_project, s[i])))
    project_out = dict(zip(file_project, t))
    return project_out


def walk_dbsta(scan_dir):
    for root, dirs_walk, files in os.walk(scan_dir):
        for d in dirs_walk:
            if d.lower() in dir_filters:
                key = d.lower()
                dir_walk = os.path.join(root, d)
                for f in os.listdir(dir_walk):
                    if os.path.isfile(
                            os.path.join(
                                dir_walk,
                                f)) and dir_filters[key].match(
                        f.lower()):
                        yield (key, dir_walk, f)
    return


def get_list_sheet(work_path_project):
    wb = load_workbook(work_path_project)
    # list = wb.sheetnames
    list_station = list(filter(lambda x: not (x.startswith('К_') or x.startswith('ТУ_')), wb.sheetnames))
    # list_station = []
    # for l in list:
    #     if not (l.startswith('К_') or l.startswith('ТУ_')):
    #         list_station.append(l)
    return list_station
