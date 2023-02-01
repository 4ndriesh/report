import re
import os
from openpyxl import load_workbook, Workbook

dir_filters = {
    'project': re.compile('.*(\.xlsx)$'),
    '3': re.compile('.*(\.xlsx)$'),
    '44': re.compile('.*(\.xlsx)$')
    # 'report': re.compile('.*(\.xlsx)$')
}

def walk_dbsta(scan_dir):
    for root, dirs_walk, files in os.walk(scan_dir):
        for d in dirs_walk:
            if d.lower() in dir_filters:
                key = d.lower()
                dir_walk = os.path.join(root, d)
                for f in os.listdir(dir_walk):
                    if os.path.isfile(
                            os.path.join(
                                dir_walk,
                                f)) and dir_filters[key].match(
                        f.lower()):
                        yield (key, dir_walk, f)
    return


def get_list_sheet(work_path_project):
    wb = load_workbook(work_path_project)
    list_station = list(filter(lambda x: not (x.startswith('К_') or x.startswith('ТУ_')), wb.sheetnames))
    return list_station
