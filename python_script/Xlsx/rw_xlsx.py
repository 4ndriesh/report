# -*- coding: utf-8 -*-
# Чтение файла шаблона xlsx для получение параметров
import shutil

from functools import wraps
import pandas as pd
from openpyxl import load_workbook
from python_script import *
# from pathlib import Path

from python_script.Logging import *
from basesetting import BaseSettings

dir = BaseSettings()
log = Logger()


@singleton
class Xlsx:
    def __init__(self):
        pass
    def validate_args(func):
        @wraps(func)
        def wrapped(*args, **kwargs):
            path = kwargs.get('path')
            if (path.exists()):
                log.message_info(path)                                
                return func(*args, **kwargs)
            else:
                log.message_error('{0}{1}'.format('Проверить ', path))                                
                return pd.DataFrame()
                # raise Exception('{0}{1}'.format('Проверить ', path))
        return wrapped

    @validate_args
    def open_sheet(self, sheet_name=None, skiprows=7,path=''):
        dict_Sheet=pd.read_excel(path, sheet_name=sheet_name, skiprows=skiprows)
        return dict_Sheet

    @validate_args
    def open_book(self,path=''):
        wb=load_workbook(path, read_only=True)
        listsheets = wb.sheetnames
        wb.close()
        return listsheets

    def copy_xlsx(self,path_temp,path_temp_to):
        shutil.copy(path_temp, path_temp_to)

    @validate_args
    def read_csv(self,path='',skiprows=1,delimiter=';'):
            return pd.read_csv(path, encoding='cp1251', delimiter=delimiter, skiprows=skiprows)

    @validate_args
    def write_to_excel(self, data_for_record,startrow, index, conv_file,path=''):
        mv = load_workbook(path,keep_vba=True)
        with pd.ExcelWriter(path, engine='openpyxl') as write_to_report:
            write_to_report._book = mv
            write_to_report._sheets = dict((ws.title, ws) for ws in mv.worksheets)
            for Sheet in data_for_record.keys():
                data_for_record.get(Sheet).to_excel(write_to_report,
                                                         sheet_name=Sheet,
                                                         startrow=startrow,
                                                         startcol=0,
                                                         header=False,
                                                         index=index)
                sh = mv[Sheet] 
                if(Sheet=='ТУ' or 'ТС'):
                    sh.oddFooter.left.text=dir.CONFIG.get('Position')    
                    sh.oddFooter.center.text=dir.CONFIG.get('Date')    
                    sh.oddFooter.right.text= dir.CONFIG.get('Name')               
                sh.cell(row=4, column=1, value='{0}{1}'.format('Станция ', conv_file))
            write_to_report._save()